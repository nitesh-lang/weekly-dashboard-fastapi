"""
Build BSR snapshots JSON from all CSVs in data/BSR/<brand>/*.csv
Reads every dated Keepa CSV, groups by detected date, writes to
src/bsr_snapshots.json which gets bundled into the app.

Output format (same as old localStorage):
{
  "2026-04-23": {
    "Audio Array":    [ { asin, image, title, bsr, rating, reviews, ... }, ... ],
    "Nexlev":         [ ... ],
    "Tonor":          [ ... ],
    "White Mulberry": [ ... ]
  },
  "2026-04-24": { ... }
}

Run: python scripts/build_bsr_snapshots.py
"""
import csv
import json
import os
import re
import sys
from pathlib import Path

# ── Path setup ─────────────────────────────────────────────────────────
# Script can be run from anywhere; it anchors to project root (parent of scripts/)
SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
BSR_DIR  = PROJECT_ROOT / "data" / "BSR"
OUT_PATH = PROJECT_ROOT / "src" / "bsr_snapshots.json"

# ── Brand detection ────────────────────────────────────────────────────
BRANDS = {
    "audio":    "Audio Array",
    "nexlev":   "Nexlev",
    "tonor":    "Tonor",
    "mulberry": "White Mulberry",
    "white":    "White Mulberry",
}

def detect_brand(path_parts):
    """Look at folder name and filename for brand keywords."""
    text = " ".join(path_parts).lower()
    for key, brand in BRANDS.items():
        if key in text:
            return brand
    return None

# ── Date parsing ───────────────────────────────────────────────────────
def extract_date(filename):
    """Parse date from filename. Returns 'YYYY-MM-DD' or None."""
    name = re.sub(r"\.(csv|xlsx|xlsm)$", "", filename, flags=re.IGNORECASE)
    # YYYY-MM-DD
    m = re.search(r"(\d{4})[-_](\d{2})[-_](\d{2})", name)
    if m:
        y, mo, d = m.groups()
        if 1 <= int(mo) <= 12 and 1 <= int(d) <= 31:
            return f"{y}-{mo}-{d}"
    # DD-MM-YYYY
    m = re.search(r"(\d{2})[-_](\d{2})[-_](\d{4})", name)
    if m:
        d, mo, y = m.groups()
        if 1 <= int(mo) <= 12 and 1 <= int(d) <= 31:
            return f"{y}-{mo}-{d}"
    # YY-MM-DD (2-digit year assumed 20xx)
    m = re.search(r"(?:^|[^0-9])(\d{2})[-_](\d{2})[-_](\d{2})(?:$|[^0-9])", name)
    if m:
        y, mo, d = m.groups()
        if 1 <= int(mo) <= 12 and 1 <= int(d) <= 31:
            return f"20{y}-{mo}-{d}"
    return None

# ── CSV parsing — mirrors the parseKeepaCsv() logic in the React app ───
def to_num(v):
    if v is None: return None
    s = str(v).replace(",", "").replace("#", "").replace("%", "").strip()
    if not s or s in ("-", "—", "N/A", "n/a"): return None
    try: return float(s)
    except ValueError: return None

def read_table(filepath):
    """Read a .csv or .xlsx export into (fieldnames, [row dicts]).

    Keepa exports arrive in both shapes -- the scheduled fetch writes CSV, a
    hand-made export from the Keepa site is usually XLSX -- and the rest of this
    script should not care which one it got.
    """
    if filepath.suffix.lower() in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook

        # read_only keeps a big export from being loaded into memory twice;
        # data_only takes the cached value of a formula rather than its text.
        wb = load_workbook(filepath, read_only=True, data_only=True)
        try:
            ws = wb.active
            it = ws.iter_rows(values_only=True)
            header = next(it, None)
            if header is None:
                return [], []
            fieldnames = [("" if h is None else str(h).strip()) for h in header]
            rows = []
            for values in it:
                # A trailing blank row is common in hand-saved sheets.
                if values is None or all(v is None or str(v).strip() == "" for v in values):
                    continue
                row = {}
                for i, name in enumerate(fieldnames):
                    if not name:
                        continue
                    v = values[i] if i < len(values) else None
                    row[name] = "" if v is None else str(v)
                rows.append(row)
            return fieldnames, rows
        finally:
            wb.close()

    with open(filepath, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        return list(reader.fieldnames or []), list(reader)


def parse_bsr_file(filepath):
    """Parse a Keepa BSR export (.csv or .xlsx) into a list of row dicts."""
    try:
            fieldnames, raw_rows = read_table(filepath)
            headers = {h.strip(): h for h in fieldnames}
            # Find column names (Keepa varies them slightly)
            def col(*alts):
                for a in alts:
                    for k in headers:
                        if a.lower() == k.lower().strip():
                            return headers[k]
                # Fuzzy match
                for a in alts:
                    for k in headers:
                        if a.lower() in k.lower():
                            return headers[k]
                return None

            asin_col   = col("ASIN")
            title_col  = col("Title")
            image_col  = col("Image", "Images")
            bsr_col    = col("Sales Rank: Current", "Current Rank")
            bsr30_col  = col("Sales Rank: 30 days avg.", "30 days avg.")
            bsr90_col  = col("Sales Rank: 90 days avg.", "90 days avg.")
            bsr365_col = col("Sales Rank: 365 days avg.", "365 days avg.")
            rating_col = col("Reviews: Rating", "Rating")
            rc_col     = col("Reviews: Rating Count", "Rating Count", "Review Count")
            drops_col  = col("Sales Rank: Drops last 30 days", "30 days drop %", "Drops")
            sold_col   = col("Monthly Sales Trends: Monthly Sold (Last Known)", "Monthly Sold")

            rows = []
            for r in raw_rows:
                asin = str(r.get(asin_col, "") if asin_col else "").strip()
                if not asin:
                    continue
                image = r.get(image_col, "") if image_col else ""
                # Image column may contain multiple URLs separated by ';'
                if image:
                    image = str(image).split(";")[0].strip()
                row = {
                    "asin":   asin,
                    "image":  image or None,
                    "title":  str(r.get(title_col, "") if title_col else "").strip() or None,
                    "bsr":    to_num(r.get(bsr_col, "")) if bsr_col else None,
                    "bsr_30d":  to_num(r.get(bsr30_col, ""))  if bsr30_col  else None,
                    "bsr_90d":  to_num(r.get(bsr90_col, ""))  if bsr90_col  else None,
                    "bsr_365d": to_num(r.get(bsr365_col, "")) if bsr365_col else None,
                    "rating":  to_num(r.get(rating_col, "")) if rating_col else None,
                    "reviews": to_num(r.get(rc_col, ""))     if rc_col     else None,
                    "drops_30d":    to_num(r.get(drops_col, "")) if drops_col else None,
                    "monthly_sold": to_num(r.get(sold_col, ""))  if sold_col  else None,
                }
                rows.append(row)
            return rows
    except Exception as e:
        print(f"  ⚠️  Error reading {filepath}: {e}", file=sys.stderr)
        return []


# ── Carry-forward ──────────────────────────────────────────────────────
# Only what IDENTIFIES a product is carried forward -- its picture and its name.
# The 2026-08-05 upload had an Image column that was empty for all 530 rows, and
# blanking every product picture would make the tracker unreadable even though
# nothing about the products had changed.
#
# MEASUREMENTS ARE NEVER CARRIED FORWARD. Sales rank, rating, review count and
# monthly sold show what the export actually contained, and stay blank when it
# contained nothing. A stale rank presented as today's is worse than no rank:
# the whole point of this tracker is to say where a product ranks NOW.
CARRY_FIELDS = ("image", "title")

def carry_forward(snapshots):
    """Fill blank CARRY_FIELDS from the newest earlier snapshot of that ASIN."""
    last_known = {}   # {(brand, asin): {field: value}}
    filled = {f: 0 for f in CARRY_FIELDS}

    for date in sorted(snapshots):
        for brand, rows in snapshots[date].items():
            for row in rows:
                key = (brand, row["asin"])
                seen = last_known.setdefault(key, {})
                for f in CARRY_FIELDS:
                    if row.get(f) not in (None, ""):
                        seen[f] = row[f]          # fresh value wins, and is remembered
                    elif f in seen:
                        row[f] = seen[f]          # blank inherits the last known one
                        filled[f] += 1
    return filled

# ── Main ───────────────────────────────────────────────────────────────
def main():
    if not BSR_DIR.exists():
        print(f"❌ BSR directory not found: {BSR_DIR}")
        sys.exit(1)

    # Collect all CSVs with their detected brand + date
    files_by_key = {}  # {(date, brand): (path, has_images_count)}
    total_scanned = 0

    for brand_folder in sorted(BSR_DIR.iterdir()):
        if not brand_folder.is_dir():
            continue
        data_files = sorted(
            p for p in brand_folder.iterdir()
            if p.is_file() and p.suffix.lower() in (".csv", ".xlsx", ".xlsm")
            and not p.name.startswith("~$")   # Excel's lock files
        )
        for data_file in data_files:
            total_scanned += 1
            date = extract_date(data_file.name)
            brand = detect_brand([brand_folder.name, data_file.name])
            if not date:
                print(f"  ⏩ Skipping {data_file.name} — no date in filename")
                continue
            if not brand:
                print(f"  ⏩ Skipping {data_file} — no brand detected")
                continue

            # Dedup: if the same (date, brand) has more than one file, prefer
            # the one carrying more rows, and break ties on images. Row count
            # leads because a fuller export is the better record; the image gap
            # is repaired afterwards by carry_forward().
            key = (date, brand)
            rows = parse_bsr_file(data_file)
            has_img = sum(1 for r in rows if r.get("image"))
            if key in files_by_key:
                _, prev_rows, prev_has_img = files_by_key[key]
                if (len(rows), has_img) <= (len(prev_rows), prev_has_img):
                    continue  # keep the existing one
            files_by_key[key] = (data_file, rows, has_img)

    # Build output: { date: { brand: [rows] } }
    snapshots = {}
    for (date, brand), (path, rows, has_img) in sorted(files_by_key.items()):
        snapshots.setdefault(date, {})[brand] = rows
        print(f"  ✓ {date}  {brand:16s}  {len(rows):4d} ASINs  {has_img:4d} images  ({path.name})")

    filled = carry_forward(snapshots)
    if any(filled.values()):
        detail = ", ".join(f"{n} {f}" for f, n in filled.items() if n)
        print(f"\n🔁 Carried forward from earlier days: {detail}")

    # Summary
    dates = sorted(snapshots.keys())
    total_asins = sum(len(rows) for d in snapshots.values() for rows in d.values())
    print(f"\n📊 Scanned {total_scanned} data files")
    print(f"   → {len(dates)} unique date{'s' if len(dates)!=1 else ''}: {', '.join(dates) if dates else '(none)'}")
    print(f"   → {total_asins} total ASIN rows")

    # Write output
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text(json.dumps(snapshots, separators=(",", ":"), ensure_ascii=False), encoding="utf-8")
    size_kb = OUT_PATH.stat().st_size / 1024
    print(f"\n✅ Wrote {OUT_PATH.relative_to(PROJECT_ROOT)}  ({size_kb:.1f} KB)")

if __name__ == "__main__":
    main()

