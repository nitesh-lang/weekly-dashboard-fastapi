"""
Normalize Week 21 inventory files in-place:

  1. Model casing — for any inv row whose SKU is in master and whose Model
     matches the master Model case-insensitively but differs in casing,
     rewrite Model to the master's casing.  Rows where master and inv
     genuinely disagree on the model (e.g. FBA80125 AA-26 Pro vs AA-28)
     are NOT touched — those are real conflicts for manual resolution.

  2. Leading/trailing whitespace — on SKU / ASIN / Brand / Model cells,
     strip whitespace and collapse internal runs of spaces to a single
     space.  Catches "Nexlev " etc.

Saves a .bak copy of each modified workbook next to the original.
"""
from __future__ import annotations
from pathlib import Path
import shutil

import pandas as pd
from openpyxl import load_workbook

ROOT   = Path(__file__).resolve().parent.parent
MASTER = ROOT / "data" / "master" / "sku_master.xlsx"
WEEK21 = ROOT / "data" / "raw" / "inventory" / "Week 21"
BRANDS = ["Audio_Array", "Nexlev", "Tonor", "White_Mulberry"]


def _norm(s) -> str:
    return "" if pd.isna(s) else str(s).strip()


# ── Master: SKU → canonical Model ────────────────────────────────────────
m = pd.read_excel(MASTER)
m.columns = m.columns.str.strip()
sku_to_model: dict[str, str] = {}
for _, r in m.iterrows():
    sku   = _norm(r.get("FBA SKU"))
    model = _norm(r.get("Model"))
    if sku and model:
        sku_to_model[sku] = model

print(f"Master: {len(sku_to_model):,} SKU→Model entries\n")

total_changes = 0
for brand_dir in BRANDS:
    p = WEEK21 / brand_dir / "Inventory Snapshot.xlsx"
    if not p.exists():
        continue

    wb = load_workbook(p)
    ws = wb.active

    # Locate identifier columns (header row = 1)
    header = {str(c.value).strip(): c.column for c in ws[1] if c.value}
    if "SKU" not in header or "Model" not in header:
        print(f"⚠ {brand_dir}: missing SKU/Model header, skipping")
        continue
    sku_col   = header["SKU"]
    model_col = header["Model"]
    id_cols   = {c: header[c] for c in ("SKU", "ASIN", "Brand", "Model") if c in header}

    casing_changes: list[tuple[int, str, str, str]] = []   # (row, sku, before, after)
    ws_changes:     list[tuple[int, str, str, str]] = []   # (row, col, before, after)

    for row in range(2, ws.max_row + 1):
        # ── Whitespace strip on all identifier columns ──
        for col_name, col_idx in id_cols.items():
            cell = ws.cell(row=row, column=col_idx)
            if cell.value is None:
                continue
            raw = str(cell.value)
            cleaned = " ".join(raw.split())   # strip + collapse internal runs
            if cleaned != raw:
                cell.value = cleaned if cleaned else None
                ws_changes.append((row, col_name, raw, cleaned))

        # ── Model casing fix against master ──
        sku_cell   = ws.cell(row=row, column=sku_col)
        model_cell = ws.cell(row=row, column=model_col)
        sku   = _norm(sku_cell.value)
        model = _norm(model_cell.value)
        if not sku or not model:
            continue
        canon = sku_to_model.get(sku)
        if not canon:
            continue
        if canon != model and canon.lower() == model.lower():
            model_cell.value = canon
            casing_changes.append((row, sku, model, canon))

    if not casing_changes and not ws_changes:
        print(f"  {brand_dir.replace('_',' ')}: nothing to fix")
        continue

    # Backup before writing
    bak = p.with_suffix(p.suffix + ".bak")
    if not bak.exists():
        shutil.copy2(p, bak)
        print(f"  💾 Backup: {bak.relative_to(ROOT)}")

    wb.save(p)
    total_changes += len(casing_changes) + len(ws_changes)
    print(f"  {brand_dir.replace('_',' ')}: "
          f"{len(casing_changes)} casing · {len(ws_changes)} whitespace")
    for row, sku, before, after in casing_changes[:10]:
        print(f"      casing   row {row:>4}  SKU {sku!r:<14}  {before!r:<22} → {after!r}")
    if len(casing_changes) > 10:
        print(f"      ...and {len(casing_changes)-10} more casing fixes")
    for row, col, before, after in ws_changes[:10]:
        print(f"      ws-strip row {row:>4}  {col:<6}  {before!r:<22} → {after!r}")
    if len(ws_changes) > 10:
        print(f"      ...and {len(ws_changes)-10} more ws fixes")

print(f"\n✅ Total: {total_changes} cell(s) updated across {len(BRANDS)} files")
