"""
Produce a step-by-step action list for every raw-row exception the
strict-ASIN sync left untouched.  One Excel file with per-row
"open this file → go to this sheet → row N → change X to Y" instructions.

Output: data/processed/raw_exceptions_action_list.xlsx
"""
from __future__ import annotations
from pathlib import Path
import re

import pandas as pd
from openpyxl import load_workbook

ROOT       = Path(__file__).resolve().parent.parent
MASTER     = ROOT / "data" / "master" / "sku_master.xlsx"
SALES_ROOT = ROOT / "data" / "raw" / "sales"
AMS_ROOT   = ROOT / "data" / "ams_weekly_data"
OUT        = ROOT / "data" / "processed" / "raw_exceptions_action_list.xlsx"
BRAND_DIRS = ["Audio_Array", "Nexlev", "Tonor", "White_Mulberry"]


def _norm(s) -> str:
    return "" if s is None or (isinstance(s, float) and pd.isna(s)) else str(s).strip()


# Master ASIN lookup
m = pd.read_excel(MASTER)
m.columns = m.columns.str.strip()
master_by_asin: dict[str, dict] = {}
for _, r in m.iterrows():
    p = _norm(r.get("ASIN"))
    rec = {"sku": _norm(r.get("FBA SKU")), "model": _norm(r.get("Model")),
           "brand": _norm(r.get("Brand"))}
    if p:
        master_by_asin[p] = rec
    v = _norm(r.get("Variation ASINs"))
    if v:
        for x in re.split(r"[,\s/|;]+", v):
            x = x.strip()
            if x and x not in master_by_asin:
                master_by_asin[x] = rec

# Master SKU lookup (used only to enrich the action hint — NOT for sync)
master_by_sku: dict[str, dict] = {}
for _, r in m.iterrows():
    for col in ("FBA SKU", "Original SKU"):
        v = _norm(r.get(col))
        if v:
            master_by_sku.setdefault(v, {"asin": _norm(r.get("ASIN")),
                                          "model": _norm(r.get("Model")),
                                          "brand": _norm(r.get("Brand"))})


actions: list[dict] = []

def _record(file_rel: str, sheet: str, row: int, kind: str,
            asin: str, sku: str, model: str,
            qty: str = "", sales: str = ""):
    # Decide the action recommendation per row
    if kind == "orphan":
        # ASIN in raw but not in master
        # Is this a formula leak?  An obvious SKU-in-ASIN typo?  Or a real new ASIN?
        if asin.startswith("="):
            action = "REPLACE FORMULA — paste the correct ASIN over the formula"
            target = "ASIN cell"
            change_from = asin[:60] + ("…" if len(asin) > 60 else "")
            change_to   = "the real ASIN (look up in Seller Central)"
            why         = "Broken external XLOOKUP — references a workbook that isn't loaded, ASIN cell shows the formula text"
        elif asin.startswith("FB") and any(c.isdigit() for c in asin):
            # Looks like a SKU got pasted into the ASIN cell
            action = "FIX TYPO — the ASIN cell has a SKU value"
            target = "ASIN cell"
            change_from = asin
            mr = master_by_sku.get(asin)
            change_to = (f"the real ASIN for SKU {asin}"
                         + (f" (master says {mr['asin']})" if mr and mr.get("asin") else ""))
            why = "ASIN cell contains a SKU pattern — operator pasted in the wrong column"
        else:
            # A real ASIN, just not in master yet
            action = "ADD TO MASTER — paste this ASIN as a new row"
            target = "sku_master.xlsx · ASIN column"
            change_from = "(missing)"
            change_to = f"ASIN={asin}, SKU={sku or '?'}, Model={model or '?'}"
            why = "New listing Amazon recognises but master doesn't"
    elif kind == "no_asin":
        # Row has SKU/Model but no ASIN value
        # Special case: "total" footer rows
        if not sku and model.lower() == "total":
            action = "DELETE ROW — this is a summary footer, not real data"
            target = "the entire row"
            change_from = "(footer row)"
            change_to = "(delete)"
            why = "D2C export drops a 'total' summary line at the bottom of each channel"
        else:
            # Real sale row missing ASIN
            mr = master_by_sku.get(sku) if sku else None
            if mr and mr.get("asin"):
                action = "FILL ASIN — paste master's ASIN into the ASIN cell"
                target = "ASIN cell"
                change_from = "(empty)"
                change_to = mr["asin"]
                why = f"SKU {sku} exists in master with ASIN {mr['asin']} (model {mr.get('model','?')}) — operator left ASIN blank"
            else:
                action = "FILL ASIN — look up in Seller Central and paste in"
                target = "ASIN cell"
                change_from = "(empty)"
                change_to = "the real ASIN"
                why = f"SKU {sku!r} not in master either — operator should also add this SKU to master"

    actions.append({
        "Severity":   "HIGH" if (kind == "orphan" and not asin.startswith("=") and not asin.startswith("FB")) else
                      "HIGH" if kind == "no_asin" and not (not sku and model.lower() == "total") else
                      "LOW",
        "Action":     action,
        "Target":     target,
        "File":       file_rel,
        "Sheet":      sheet,
        "Row #":      row,
        "Current":    change_from,
        "Change to":  change_to,
        "Reason":     why,
        "Raw ASIN":   asin or "(empty)",
        "Raw SKU":    sku or "(empty)",
        "Raw Model":  model or "(empty)",
        "Qty / Units":qty or "",
        "Sale Amount":sales or "",
    })


def _walk(path: Path, asin_cols: tuple[str, ...], file_rel: str):
    if not path.exists():
        return
    try:
        wb = load_workbook(path)
    except PermissionError:
        return
    for ws in wb.worksheets:
        header = {str(c.value).strip(): c.column for c in ws[1] if c.value}
        sku_col   = header.get("SKU")
        model_col = header.get("Model")
        asin_cands = [header[c] for c in asin_cols if c in header]
        qty_col   = next((header[k] for k in ("Qty","units_ordered") if k in header), None)
        sales_col = next((header[k] for k in ("Sale Amount","ordered_product_sales") if k in header), None)
        if not asin_cands or sku_col is None or model_col is None:
            continue
        for row in range(2, ws.max_row + 1):
            asin = ""
            for ci in asin_cands:
                v = _norm(ws.cell(row=row, column=ci).value)
                if v: asin = v; break
            sku   = _norm(ws.cell(row=row, column=sku_col).value)
            model = _norm(ws.cell(row=row, column=model_col).value)
            qty   = _norm(ws.cell(row=row, column=qty_col).value)   if qty_col   else ""
            sales = _norm(ws.cell(row=row, column=sales_col).value) if sales_col else ""

            if not asin:
                # Skip totally blank trailing rows
                if not sku and not model:
                    continue
                _record(file_rel, ws.title, row, "no_asin", asin, sku, model, qty, sales)
                continue

            if asin not in master_by_asin:
                _record(file_rel, ws.title, row, "orphan", asin, sku, model, qty, sales)
                continue
    wb.close()


# Walk every relevant file
for wk in range(10, 22):
    for bd in BRAND_DIRS:
        _walk(SALES_ROOT/f"Week {wk}"/bd/"amazon_sales.xlsx",
              ("(Parent) ASIN","(Child) ASIN"),
              f"data/raw/sales/Week {wk}/{bd}/amazon_sales.xlsx")
        _walk(SALES_ROOT/f"Week {wk}"/bd/"other_channels.xlsx",
              ("ASIN",),
              f"data/raw/sales/Week {wk}/{bd}/other_channels.xlsx")
        _walk(AMS_ROOT/bd/f"business_report_week{wk}.xlsx",
              ("(Parent) ASIN","(Child) ASIN"),
              f"data/ams_weekly_data/{bd}/business_report_week{wk}.xlsx")


# Write the Excel
OUT.parent.mkdir(parents=True, exist_ok=True)
df = pd.DataFrame(actions, columns=[
    "Severity","Action","Target","File","Sheet","Row #",
    "Current","Change to","Reason",
    "Raw ASIN","Raw SKU","Raw Model","Qty / Units","Sale Amount",
])
# Sort by Severity (HIGH first), then File, Sheet, Row
sev_order = {"HIGH": 0, "MEDIUM": 1, "LOW": 2}
df["_sev"] = df["Severity"].map(sev_order).fillna(3)
df = df.sort_values(["_sev","File","Sheet","Row #"]).drop(columns="_sev")

summary = (df.groupby(["Severity","Action"], as_index=False).size()
             .rename(columns={"size":"Rows"})
             .sort_values(["Severity","Rows"], ascending=[True,False]))

with pd.ExcelWriter(OUT, engine="openpyxl") as w:
    summary.to_excel(w, sheet_name="Summary",  index=False)
    df.to_excel(w,      sheet_name="Actions",  index=False)

print(f"📋 Action list: {OUT.relative_to(ROOT)}")
print(f"   Total exceptions: {len(df)}")
print()
print(summary.to_string(index=False))
