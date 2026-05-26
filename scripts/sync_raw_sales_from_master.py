"""
Sync SKU + Model columns in the raw sales files to whatever sku_master
says for the row's ASIN.  Master is the single source of truth — the
operator no longer maintains SKU / Model in the raw export.

Runs across the last N weeks (default 12) so historical snapshots stay
consistent with the current master.  When master changes (e.g. SKU
prefix updated, model renamed), past weeks' raw files would otherwise
still carry the old identifiers and the snapshot would mis-attribute
those weeks' sales.

For every row in:
  data/raw/sales/Week N/<brand>/amazon_sales.xlsx
  data/raw/sales/Week N/<brand>/other_channels.xlsx
  data/ams_weekly_data/<brand>/business_report_weekN.xlsx

…look up the (Parent) ASIN (amazon_sales / business_report) or ASIN
(other_channels) in master.  If the ASIN resolves (either as a primary
ASIN or as a known Variation ASIN), overwrite the SKU and Model cells
with master's canonical values.

STRICT ASIN-ONLY RULE (per operator decision):
  • ASIN missing in raw           → row left untouched, flagged "no-ASIN"
  • ASIN present but NOT in master → row left untouched, flagged "orphan"
  • No SKU-based fallback.  To resolve, add the ASIN to master.

Saves a one-shot `.bak` next to each modified workbook before writing.

CLI:
  python scripts/sync_raw_sales_from_master.py          # last 12 weeks
  python scripts/sync_raw_sales_from_master.py 8        # last 8 weeks
  python scripts/sync_raw_sales_from_master.py 21 21    # single week
  python scripts/sync_raw_sales_from_master.py 10 21    # explicit range
"""
from __future__ import annotations
from pathlib import Path
import re
import shutil
import sys

import pandas as pd
from openpyxl import load_workbook

ROOT        = Path(__file__).resolve().parent.parent
MASTER      = ROOT / "data" / "master" / "sku_master.xlsx"
SALES_ROOT  = ROOT / "data" / "raw" / "sales"
AMS_ROOT    = ROOT / "data" / "ams_weekly_data"   # business_report_week<N>.xlsx lives here
BRAND_DIRS  = ["Audio_Array", "Nexlev", "Tonor", "White_Mulberry"]   # no Fossil
DEFAULT_WINDOW = 12


def _norm(s) -> str:
    return "" if s is None or (isinstance(s, float) and pd.isna(s)) else str(s).strip()


# ── Master: build ASIN → canonical (sku, model) ──────────────────────────
m = pd.read_excel(MASTER)
m.columns = m.columns.str.strip()
for c in ("FBA SKU", "ASIN", "Model", "Variation ASINs"):
    if c in m.columns:
        m[c] = m[c].map(_norm)

master_by_asin: dict[str, dict] = {}
for _, r in m.iterrows():
    primary = r.get("ASIN", "")
    rec = {"sku": r.get("FBA SKU", ""), "model": r.get("Model", "")}
    if primary:
        master_by_asin[primary] = rec
    vstr = r.get("Variation ASINs", "")
    if vstr:
        for v in re.split(r"[,\s/|;]+", vstr):
            v = v.strip()
            if v and v not in master_by_asin:
                master_by_asin[v] = rec

print(f"Master ASINs (incl. variations): {len(master_by_asin):,}\n")


# ── In-place sync via openpyxl (preserves formatting) ────────────────────
#
# Strict ASIN-only rule (per operator):
#   • ASIN in raw + in master → overwrite SKU + Model from master
#   • ASIN in raw + NOT in master → orphan; row left as-is
#   • ASIN missing in raw       → no-ASIN; row left as-is
# No SKU-based fallback. If the operator wants those rows resolved,
# they must add the ASIN to master.
def _sync_workbook(path: Path, asin_cols: tuple[str, ...]) -> dict:
    """Returns {cells, orphans, missing_asin, orphan_log, missing_log}."""
    blank = {"cells": 0, "orphans": 0, "missing_asin": 0,
             "orphan_log": [], "missing_log": []}
    if not path.exists():
        return blank
    try:
        wb = load_workbook(path)
    except PermissionError:
        print(f"   ⚠ {path.name} is open in Excel — close it and re-run.")
        return blank

    # `other_channels.xlsx` files carry one sheet per channel (CRED_B2C,
    # Pharmaeasy, Myntra, …).  We iterate every sheet that has the
    # expected schema so all channels get the master-as-truth treatment,
    # not just whichever one happens to be `wb.active`.
    cell_updates = 0
    orphan_log   = []
    missing_log  = []

    sheets_processed = 0
    for ws in wb.worksheets:
        header = {str(c.value).strip(): c.column for c in ws[1] if c.value}
        sku_col   = header.get("SKU")
        model_col = header.get("Model")
        asin_col_candidates = [header[c] for c in asin_cols if c in header]
        if not asin_col_candidates or sku_col is None or model_col is None:
            # Sheet doesn't carry our schema — skip silently (only warn
            # when no sheet at all is usable, handled below)
            continue
        sheets_processed += 1
        if ws.max_row < 2:
            continue
        # Channel label used in the per-row log so multi-sheet files don't
        # become ambiguous.
        sheet_label = ws.title

        for row in range(2, ws.max_row + 1):
            # Use the first ASIN column that has a value (parent first, then child)
            asin = ""
            for col_idx in asin_col_candidates:
                v = _norm(ws.cell(row=row, column=col_idx).value)
                if v:
                    asin = v
                    break

            if not asin:
                # No ASIN in raw — strict rule: leave row untouched.
                raw_sku   = _norm(ws.cell(row=row, column=sku_col).value)
                raw_model = _norm(ws.cell(row=row, column=model_col).value)
                if raw_sku or raw_model:
                    missing_log.append(f"[{sheet_label}] row {row} · SKU={raw_sku or '—'} · Model={raw_model or '—'}")
                continue

            mrec = master_by_asin.get(asin)
            if mrec is None:
                # ASIN not in master — strict rule: leave row untouched.
                orphan_log.append(f"[{sheet_label}] row {row} · ASIN {asin}")
                continue

            # Overwrite SKU if differs (only when master has a SKU on file)
            if mrec["sku"]:
                cell = ws.cell(row=row, column=sku_col)
                cur  = _norm(cell.value)
                if cur != mrec["sku"]:
                    cell.value = mrec["sku"]
                    cell_updates += 1
            # Overwrite Model if differs (only when master has a Model on file)
            if mrec["model"]:
                cell = ws.cell(row=row, column=model_col)
                cur  = _norm(cell.value)
                if cur != mrec["model"]:
                    cell.value = mrec["model"]
                    cell_updates += 1

    if sheets_processed == 0:
        print(f"   ⚠ {path.name}: no usable sheet (missing SKU/Model/ASIN headers)")
        return blank

    if cell_updates:
        bak = path.with_suffix(path.suffix + ".bak")
        if not bak.exists():
            shutil.copy2(path, bak)
            print(f"   💾 backup → {bak.relative_to(ROOT)}")
        try:
            wb.save(path)
        except PermissionError:
            print(f"   ⚠ couldn't save {path.name} — close it in Excel and re-run.")
            return {"cells": 0,
                    "orphans": len(orphan_log), "missing_asin": len(missing_log),
                    "orphan_log": orphan_log, "missing_log": missing_log}

    return {"cells": cell_updates,
            "orphans": len(orphan_log), "missing_asin": len(missing_log),
            "orphan_log": orphan_log, "missing_log": missing_log}


# ── Resolve week range from CLI ──────────────────────────────────────────
def _wknum(name: str) -> int:
    try:    return int(name.replace("Week", "").strip())
    except: return -1

available = sorted(
    [(_wknum(p.name), p) for p in SALES_ROOT.iterdir() if p.is_dir() and p.name.startswith("Week ")],
    key=lambda x: x[0],
)
all_weeks = [w for w, _ in available if w >= 0]

if len(sys.argv) == 3:
    lo, hi = int(sys.argv[1]), int(sys.argv[2])
    weeks_in_scope = [w for w in all_weeks if lo <= w <= hi]
elif len(sys.argv) == 2:
    n = int(sys.argv[1])
    weeks_in_scope = all_weeks[-n:]
else:
    weeks_in_scope = all_weeks[-DEFAULT_WINDOW:]

print(f"Scope: weeks {weeks_in_scope[0]}–{weeks_in_scope[-1]}  ({len(weeks_in_scope)} weeks)\n")


# ── Walk every week × brand ──────────────────────────────────────────────
total_cells     = 0
total_orphan    = 0
total_missing   = 0
orphan_report:  list[dict] = []
missing_report: list[dict] = []
per_week_summary: list[dict] = []

for wk in weeks_in_scope:
    wk_dir = SALES_ROOT / f"Week {wk}"
    if not wk_dir.exists():
        continue
    print(f"━━━ Week {wk} ━━━")
    wk_cells = 0
    wk_orph  = 0
    wk_miss  = 0

    for bdir in BRAND_DIRS:
        brand_dir = wk_dir / bdir
        if not brand_dir.exists():
            continue

        def _record(file_label: str, res: dict):
            for ln in res["orphan_log"]:
                orphan_report.append({"Week": wk, "File": f"{bdir}/{file_label}", "Detail": ln})
            for ln in res["missing_log"]:
                missing_report.append({"Week": wk, "File": f"{bdir}/{file_label}", "Detail": ln})

        p = brand_dir / "amazon_sales.xlsx"
        if p.exists():
            res = _sync_workbook(p, asin_cols=("(Parent) ASIN", "(Child) ASIN"))
            if res["cells"] or res["orphans"] or res["missing_asin"]:
                print(f"   [{bdir.replace('_',' '):<14}] amazon_sales.xlsx     "
                      f"· {res['cells']} cell(s) · {res['orphans']} orphan(s) · {res['missing_asin']} no-ASIN")
            wk_cells += res["cells"]; wk_orph += res["orphans"]; wk_miss += res["missing_asin"]
            _record("amazon_sales.xlsx", res)

        p = brand_dir / "other_channels.xlsx"
        if p.exists():
            res = _sync_workbook(p, asin_cols=("ASIN",))
            if res["cells"] or res["orphans"] or res["missing_asin"]:
                print(f"   [{bdir.replace('_',' '):<14}] other_channels.xlsx   "
                      f"· {res['cells']} cell(s) · {res['orphans']} orphan(s) · {res['missing_asin']} no-ASIN")
            wk_cells += res["cells"]; wk_orph += res["orphans"]; wk_miss += res["missing_asin"]
            _record("other_channels.xlsx", res)

    # ── Business reports (data/ams_weekly_data/<brand>/business_report_weekN.xlsx) ──
    # Same Seller-Central schema as amazon_sales: SKU / Model / (Parent) ASIN / (Child) ASIN.
    for bdir in BRAND_DIRS:
        p = AMS_ROOT / bdir / f"business_report_week{wk}.xlsx"
        if not p.exists():
            continue
        res = _sync_workbook(p, asin_cols=("(Parent) ASIN", "(Child) ASIN"))
        if res["cells"] or res["orphans"] or res["missing_asin"]:
            print(f"   [{bdir.replace('_',' '):<14}] business_report_week{wk}.xlsx "
                  f"· {res['cells']} cell(s) · {res['orphans']} orphan(s) · {res['missing_asin']} no-ASIN")
        wk_cells += res["cells"]; wk_orph += res["orphans"]; wk_miss += res["missing_asin"]
        for ln in res["orphan_log"]:
            orphan_report.append({"Week": wk, "File": f"{bdir}/business_report_week{wk}.xlsx", "Detail": ln})
        for ln in res["missing_log"]:
            missing_report.append({"Week": wk, "File": f"{bdir}/business_report_week{wk}.xlsx", "Detail": ln})

    if wk_cells == 0 and wk_orph == 0 and wk_miss == 0:
        print(f"   (nothing to fix)")
    print()
    total_cells  += wk_cells
    total_orphan += wk_orph
    total_missing+= wk_miss
    per_week_summary.append({"Week": wk, "Cells updated": wk_cells,
                              "Orphan rows": wk_orph, "No-ASIN rows": wk_miss})

print("=" * 70)
print(f"✅ {total_cells} cell(s) updated via ASIN lookup")
print(f"   across {len(weeks_in_scope)} weeks · 4 brands · 3 file types")
print(f"   Strict ASIN-only rule: rows without a valid master ASIN are left untouched.")
print()
print("Per-week summary:")
for s in per_week_summary:
    print(f"   Week {s['Week']:<3}  cells={s['Cells updated']:<4}"
          f"  orphans={s['Orphan rows']}  no-ASIN={s['No-ASIN rows']}")
print()

if total_orphan:
    seen = set()
    print(f"⚠  Orphans (ASIN present but NOT in master) — unique list ({total_orphan} total rows):")
    for o in orphan_report:
        key = o["Detail"].split("ASIN ")[-1] if "ASIN " in o["Detail"] else o["Detail"]
        if key in seen:
            continue
        seen.add(key)
        print(f"     ASIN {key:<14}  first seen → Week {o['Week']} · {o['File']}")
    print(f"   → Add the ASIN to master, OR correct the ASIN in the raw file.")
    print()

if total_missing:
    print(f"⚠  No-ASIN rows ({total_missing} total) — raw row has SKU/Model but no ASIN to look up:")
    # Aggregate by SKU so the operator sees one line per unfindable SKU
    by_sku: dict[str, list[str]] = {}
    for r in missing_report:
        # Detail looks like "row N · SKU=X · Model=Y"
        sku_part = r["Detail"].split("SKU=")[-1].split(" · ")[0]
        by_sku.setdefault(sku_part, []).append(f"Week {r['Week']} · {r['File']} · {r['Detail']}")
    for sku in sorted(by_sku):
        instances = by_sku[sku]
        print(f"     SKU {sku:<18}  {len(instances)} occurrence(s)")
        for inst in instances[:3]:
            print(f"         {inst}")
        if len(instances) > 3:
            print(f"         …and {len(instances)-3} more")
    print(f"   → Fix the ASIN in the raw file so future runs can look it up.")
    print()
