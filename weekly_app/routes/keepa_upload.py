"""Keepa uploads from the browser — BSR CSVs (Buybox) + variation export.

The data of record for both targets lives IN GIT (Render's disk is wiped on
every deploy), so an upload here becomes a COMMIT on main via the GitHub
API. That commit auto-triggers the Render deploy, which rebuilds the
/buybox bundle (and refreshes data/master/keepa_variations.csv for the
Variation Performance page). Browser → commit → live, ~5 minutes, no
terminal involved.

Targets:
  POST /api/keepa-upload/bsr          — per-brand Keepa CSVs (or one ZIP of
                                        them) → buybox_src/data/BSR/<Brand>/
                                        <yy-mm-dd>.csv + regenerated
                                        src/bsr_snapshots.json
  POST /api/keepa-upload/variations   — the parent/child ASIN export →
                                        data/master/keepa_variations.csv
  GET  /api/keepa-upload/status       — latest BSR date per brand + counts

Auth: any non-viewer weekly login (admins + operators).
Requires env GITHUB_TOKEN with repo write on the weekly repo; without it
the endpoints answer 503 with instructions instead of failing weirdly.
"""
from __future__ import annotations

import base64
import io
import os
import re
import zipfile
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import List, Optional

import requests as _rq
from fastapi import APIRouter, File, HTTPException, Query, Request, UploadFile

from weekly_app.core import auth_users

router = APIRouter(prefix="/api/keepa-upload", tags=["keepa-upload"])

ROOT = Path(__file__).resolve().parent.parent.parent
BSR_DIR = ROOT / "buybox_src" / "data" / "BSR"
VARIATIONS_CSV = ROOT / "data" / "master" / "keepa_variations.csv"
PLANNING_DIR = ROOT / "sales_dash" / "data" / "planning"

# Sales-dashboard planning brands (folder names). Tonor rolls up into
# Audio Array; WM has no planning file — two brands by design.
PLANNING_BRANDS = {"nexlev": "nexlev", "audio_array": "audio_array"}
_PLANNING_NAME_RE = re.compile(
    r"^ASIN Planning file - (Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec) (\d{4})\.xlsx$")
_ASIN_RE = re.compile(r"^[A-Z0-9]{10}$")

REPO = "nitesh-lang/weekly-dashboard-fastapi"
IST = timezone(timedelta(hours=5, minutes=30))

# Keepa export filename → BSR brand folder. The operator's system names —
# "Nexlev", "Tonor", "Audio Array", "White Mulberry" — match EXACTLY first
# (case-insensitive); the fuzzy keys below catch variants like "Audio array".
EXACT_BRANDS = {
    "nexlev": "Nexlev",
    "tonor": "Tonor",
    "audio array": "Audio Array",
    "white mulberry": "White Mulberry",
}
BRAND_FOLDERS = {
    "nexlev": "Nexlev",
    "audio": "Audio Array",
    "tonor": "Tonor",
    "white": "White Mulberry",
    "mulberry": "White Mulberry",
}


def require_uploader(request: Request) -> str:
    email = request.session.get("user_email")
    if not email:
        raise HTTPException(401, "Login required")
    if auth_users.get_role(email) == "viewer":
        raise HTTPException(403, "Viewer role cannot upload data")
    return email


# ── GitHub single-commit helper (blobs → tree → commit → ref) ─────────────

def _gh(path: str, method: str = "GET", body: dict | None = None) -> dict:
    token = os.environ.get("GITHUB_TOKEN", "").strip()
    if not token:
        raise HTTPException(503, "GITHUB_TOKEN not configured on this service — "
                                 "set it in Render env to enable browser uploads.")
    r = _rq.request(method, f"https://api.github.com{path}",
                    headers={"Authorization": f"Bearer {token}",
                             "Accept": "application/vnd.github+json"},
                    json=body, timeout=60)
    if r.status_code >= 300:
        raise HTTPException(502, f"GitHub API {r.status_code} on {path}: {r.text[:200]}")
    return r.json()


def commit_files(files: dict[str, bytes], message: str, author_email: str) -> Optional[str]:
    """One commit on main containing `files` ({repo-relative-path: bytes}).
    Returns the commit sha, or None when nothing actually changed."""
    head = _gh(f"/repos/{REPO}/git/ref/heads/main")["object"]["sha"]
    base_commit = _gh(f"/repos/{REPO}/git/commits/{head}")

    tree_entries = []
    for path, content in files.items():
        blob = _gh(f"/repos/{REPO}/git/blobs", "POST",
                   {"content": base64.b64encode(content).decode(), "encoding": "base64"})
        tree_entries.append({"path": path.replace("\\", "/"), "mode": "100644",
                             "type": "blob", "sha": blob["sha"]})

    tree = _gh(f"/repos/{REPO}/git/trees", "POST",
               {"base_tree": base_commit["tree"]["sha"], "tree": tree_entries})
    if tree["sha"] == base_commit["tree"]["sha"]:
        return None                      # byte-identical upload — no commit

    commit = _gh(f"/repos/{REPO}/git/commits", "POST", {
        "message": message,
        "tree": tree["sha"],
        "parents": [head],
        "author": {"name": "keepa-upload (dashboard)",
                   "email": author_email or "uploads@cambiumretail.com",
                   "date": datetime.now(timezone.utc).isoformat()},
    })
    _gh(f"/repos/{REPO}/git/refs/heads/main", "PATCH",
        {"sha": commit["sha"], "force": False})
    return commit["sha"]


# ── validation helpers ────────────────────────────────────────────────────

def _brand_for(filename: str) -> Optional[str]:
    low = Path(filename).stem.strip().lower()
    if low in EXACT_BRANDS:
        return EXACT_BRANDS[low]
    for key, folder in BRAND_FOLDERS.items():
        if key in low:
            return folder
    return None


def _looks_like_keepa_bsr(content: bytes) -> bool:
    head = content[:2048].decode("utf-8-sig", "replace")
    return "ASIN" in head and "Sales Rank" in head


def _collect_csvs(uploads: List[UploadFile]) -> dict[str, bytes]:
    """Flatten uploads (CSVs and/or ZIPs of CSVs) → {brand_folder: bytes}."""
    out: dict[str, bytes] = {}
    unmatched: list[str] = []
    for up in uploads:
        raw = up.file.read()
        names_blobs: list[tuple[str, bytes]] = []
        if (up.filename or "").lower().endswith(".zip") or raw[:2] == b"PK":
            with zipfile.ZipFile(io.BytesIO(raw)) as z:
                for n in z.namelist():
                    if n.lower().endswith(".csv"):
                        names_blobs.append((n, z.read(n)))
        else:
            names_blobs.append((up.filename or "upload.csv", raw))
        for name, blob in names_blobs:
            brand = _brand_for(name)
            if brand is None:
                unmatched.append(name)
                continue
            if not _looks_like_keepa_bsr(blob):
                raise HTTPException(422, f"{name}: doesn't look like a Keepa export "
                                         "(no ASIN / Sales Rank columns in header)")
            out[brand] = blob
    if unmatched:
        raise HTTPException(422, f"Couldn't map to a brand: {', '.join(unmatched[:4])}. "
                                 "Name files like Nexlev.csv / Audio array.csv / "
                                 "Tonor.csv / White Mulberry.csv")
    if not out:
        raise HTTPException(422, "No CSVs found in the upload.")
    return out


# ── endpoints ─────────────────────────────────────────────────────────────

@router.get("/status")
def status(request: Request):
    require_uploader(request)
    brands = {}
    if BSR_DIR.exists():
        for d in sorted(BSR_DIR.iterdir()):
            if d.is_dir():
                dates = sorted(p.stem for p in d.glob("*.csv"))
                brands[d.name] = dates[-1] if dates else None
    variations = None
    if VARIATIONS_CSV.exists():
        variations = {"rows": max(sum(1 for _ in VARIATIONS_CSV.open(encoding="utf-8", errors="replace")) - 1, 0)}
    planning = {}
    if PLANNING_DIR.exists():
        for bkey, folder in PLANNING_BRANDS.items():
            months = []
            d = PLANNING_DIR / folder
            if d.exists():
                for p in d.glob("ASIN Planning file - *.xlsx"):
                    mm = _PLANNING_NAME_RE.match(p.name)
                    if mm:
                        months.append((int(mm.group(2)),
                                       ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug",
                                        "Sep","Oct","Nov","Dec"].index(mm.group(1)) + 1,
                                       f"{mm.group(1)} {mm.group(2)}"))
            planning[bkey] = sorted(months)[-1][2] if months else None
    wm_1p = []
    wm_dir = BUYBOX_DATA_DIR / "White Mulberry"
    if wm_dir.exists():
        for d in wm_dir.iterdir():
            m = _MONTH_KEY_RE.match(d.name)
            if m and (d / "1Psales.csv").exists():
                wm_1p.append((int(m.group(2)), _MONTH_ABBR.index(m.group(1)) + 1, d.name))
    return {"bsr_latest": brands, "variations": variations, "planning_latest": planning,
            "wm_1p_latest": sorted(wm_1p)[-1][2] if wm_1p else None,
            "wm_1p_months": _recent_month_keys(3),
            "github_configured": bool(os.environ.get("GITHUB_TOKEN", "").strip()),
            "today": datetime.now(IST).strftime("%y-%m-%d")}


@router.post("/bsr")
def upload_bsr(request: Request,
               files: List[UploadFile] = File(...),
               date: Optional[str] = Query(None, description="yy-mm-dd; default today IST")):
    email = require_uploader(request)
    day = date or datetime.now(IST).strftime("%y-%m-%d")
    if not re.fullmatch(r"\d{2}-\d{2}-\d{2}", day):
        raise HTTPException(422, "date must be yy-mm-dd")

    per_brand = _collect_csvs(files)

    # DELIBERATELY no snapshot regeneration here: running
    # build_bsr_snapshots.py in-request read the whole 78MB BSR store and
    # OOM-killed the 512MB instance (server_failed 2026-08-31 08:01 → the
    # operator's 502). The commit below triggers a Render deploy, and the
    # BUILD step regenerates bsr_snapshots.json from ALL committed CSVs
    # (see render.yaml) with the build container's own resources.
    commit_payload = {
        f"buybox_src/data/BSR/{brand}/{day}.csv": blob
        for brand, blob in per_brand.items()
    }

    sha = commit_files(commit_payload,
                       f"data(buybox): BSR upload {day} via dashboard "
                       f"({', '.join(sorted(per_brand))}) [skip ci]", email)
    return {"ok": True, "date": day, "brands": sorted(per_brand),
            "commit": sha,
            "note": ("✓ Already up to date — this exact data is in the system; nothing to do."
                     if sha is None else
                     "Committed. Rebuilding now — after ~5 minutes RELOAD the Buybox tab "
                     "(close & reopen, or Ctrl+Shift+R) to see the new data.")}


@router.post("/planning")
def upload_planning(request: Request,
                    brand: str = Query(..., description="nexlev | audio_array"),
                    file: UploadFile = File(...)):
    """Monthly ASIN planning workbook for the Sales Dashboard.

    HIGH-STAKES FILE: the sales app's plan-aware STRICT ingest drops every
    sales row of a month whose planning file is missing or wrong, so this
    endpoint validates hard before committing and reports exactly what it
    is replacing.
    """
    email = require_uploader(request)
    bkey = brand.strip().lower()
    if bkey not in PLANNING_BRANDS:
        raise HTTPException(422, "brand must be 'nexlev' or 'audio_array'")

    fname = (file.filename or "").strip()
    m = _PLANNING_NAME_RE.match(fname)
    if not m:
        raise HTTPException(422, "Filename must be exactly "
                                 "'ASIN Planning file - <Mon> <YYYY>.xlsx' "
                                 "(e.g. 'ASIN Planning file - Sep 2026.xlsx') — "
                                 f"got {fname!r}. The sales app matches the month "
                                 "by this exact name.")
    month_label = f"{m.group(1)} {m.group(2)}"

    raw = file.file.read()
    if raw[:4] != b"PK\x03\x04":
        raise HTTPException(422, "Not an .xlsx file.")

    # Structural validation mirroring what the sales app's loader needs.
    import openpyxl
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True)
    except Exception as e:
        raise HTTPException(422, f"Workbook unreadable: {e}")
    if "Main" not in wb.sheetnames:
        raise HTTPException(422, f"Missing 'Main' sheet (found {wb.sheetnames}). "
                                 "The sales app reads the plan from 'Main'.")
    ws = wb["Main"]
    header = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(max_row=1))]
    if "asin" not in header:
        raise HTTPException(422, f"'Main' sheet has no ASIN column (headers: {header[:6]}).")
    ai = header.index("asin")
    asins = [str(r[ai].value).strip().upper() for r in ws.iter_rows(min_row=2)
             if r[ai].value]
    valid = [a for a in asins if _ASIN_RE.match(a)]
    if len(valid) < 5:
        raise HTTPException(422, f"Only {len(valid)} valid ASINs in 'Main' — refusing: "
                                 "a near-empty plan would drop the month's sales at ingest.")
    warnings = []
    if "Category" not in wb.sheetnames:
        warnings.append("No 'Category' sheet — category targets will be empty for this month.")

    rel = f"sales_dash/data/planning/{PLANNING_BRANDS[bkey]}/{fname}"
    existing = ROOT / rel
    replaces = None
    if existing.exists():
        old = existing.read_bytes()
        if old == raw:
            return {"ok": True, "commit": None, "brand": bkey, "month": month_label,
                    "asins": len(valid), "warnings": warnings,
                    "note": "✓ Already up to date — this exact plan is in the system."}
        try:
            owb = openpyxl.load_workbook(io.BytesIO(old), read_only=True)
            ows = owb["Main"]
            oh = [str(c.value or "").strip().lower() for c in next(ows.iter_rows(max_row=1))]
            oai = oh.index("asin")
            old_n = sum(1 for r in ows.iter_rows(min_row=2) if r[oai].value)
        except Exception:
            old_n = "?"
        replaces = f"replaces existing plan ({old_n} → {len(valid)} ASINs)"

    # Commit FIRST; only mirror to local disk after the commit succeeds —
    # a failed commit must leave no half-applied state (a locally-written
    # plan with no commit would show in /status but never deploy).
    sha = commit_files({rel: raw},
                       f"data(planning): {bkey} {month_label} via dashboard"
                       + (f" — {replaces}" if replaces else "")
                       + " [skip ci]", email)
    existing.parent.mkdir(parents=True, exist_ok=True)
    existing.write_bytes(raw)
    return {"ok": True, "commit": sha, "brand": bkey, "month": month_label,
            "asins": len(valid), "replaced": bool(replaces), "warnings": warnings,
            "note": ((replaces + ". " if replaces else "")
                     + f"Committed {len(valid)} planned ASINs for {month_label}. "
                       "Live after the rebuild (~6 min). Sales for a month only land "
                       "if its plan is in BEFORE the pull — done for " + month_label + ".")}


# ── Buybox monthly — White Mulberry 1P vendor sales (manual by design) ────
# The monthly-sync workflow pulls ads + 3P for all brands and 1P for
# AA/Tonor, but the WM vendor pull is dormant — the operator downloads the
# Vendor Central "Retail Analytics" export and files it as
# buybox_src/data/White Mulberry/<MonYY>/1Psales.csv.  The Render build then
# derives raw_data.json from the committed CSVs.
BUYBOX_DATA_DIR = ROOT / "buybox_src" / "data"
_MONTH_ABBR = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
               "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
_MONTH_KEY_RE = re.compile(r"^([A-Z][a-z]{2})(\d{2})$")
_VIEWING_RE = re.compile(r"Viewing Range=\[(\d{2})/(\d{2})/(\d{2})\s*-\s*\d{2}/\d{2}/\d{2}\]")


def _recent_month_keys(n: int) -> list[str]:
    """Last n month keys as %b%y (newest first), previous month first —
    matches the pull scripts' folder naming exactly."""
    today = datetime.now(IST)
    y, m = today.year, today.month
    out = []
    for _ in range(n):
        m -= 1
        if m == 0:
            y, m = y - 1, 12
        out.append(f"{_MONTH_ABBR[m - 1]}{y % 100:02d}")
    return out


@router.post("/buybox-1p")
def upload_buybox_1p(request: Request,
                     month: str = Query(..., description="Month key, e.g. Aug26"),
                     file: UploadFile = File(...)):
    """White Mulberry 1P monthly sales for the Buybox report."""
    email = require_uploader(request)
    mk = month.strip()
    allowed = _recent_month_keys(6)
    if mk not in allowed:
        raise HTTPException(422, f"month must be one of {allowed} — got {mk!r}")

    raw = file.file.read()
    head = raw[:8192].decode("utf-8-sig", "replace")
    lines = head.splitlines()
    header_line = ""
    for ln in lines[:3]:
        if "ASIN" in ln and "Ordered Revenue" in ln:
            header_line = ln
            break
    if not header_line:
        raise HTTPException(422, "Doesn't look like the Vendor Central sales export "
                                 "(no ASIN + 'Ordered Revenue' header in the first rows). "
                                 "Upload the Retail Analytics CSV as downloaded.")

    warnings = []
    vm = _VIEWING_RE.search(head)
    if vm:
        # Banner dates are dd/mm/yy — the export's own claim of its window.
        exp_month = _MONTH_ABBR.index(mk[:3]) + 1
        exp_year = 2000 + int(mk[3:])
        got_month, got_year = int(vm.group(2)), 2000 + int(vm.group(3))
        if (got_year, got_month) != (exp_year, exp_month):
            raise HTTPException(422, f"This export's Viewing Range is "
                                     f"{vm.group(1)}/{vm.group(2)}/{vm.group(3)} — that's "
                                     f"{_MONTH_ABBR[got_month-1]} {got_year}, not {mk}. "
                                     "Pick the matching month or re-download the export.")
    else:
        warnings.append("Couldn't read the export's Viewing Range banner — "
                        "make sure this really is the " + mk + " export.")

    data_rows = max(sum(1 for ln in raw.decode("utf-8-sig", "replace").splitlines()
                        if ln.strip()) - (2 if lines and "ASIN" not in lines[0] else 1), 0)
    if data_rows < 1:
        raise HTTPException(422, "No data rows found in the export.")

    rel = f"buybox_src/data/White Mulberry/{mk}/1Psales.csv"
    existing = ROOT / rel
    if existing.exists() and existing.read_bytes() == raw:
        return {"ok": True, "commit": None, "month": mk, "rows": data_rows,
                "warnings": warnings,
                "note": "✓ Already up to date — this exact export is in the system."}

    replaced = existing.exists()
    sha = commit_files({rel: raw},
                       f"data(buybox): WM 1P sales {mk} via dashboard [skip ci]", email)
    existing.parent.mkdir(parents=True, exist_ok=True)
    existing.write_bytes(raw)
    return {"ok": True, "commit": sha, "month": mk, "rows": data_rows,
            "replaced": replaced, "warnings": warnings,
            "note": (("Replaced the existing " if replaced else "Filed ") + mk +
                     f" export ({data_rows} rows). The Buybox monthly view picks it up "
                     "after the rebuild (~6 min).")}


@router.post("/variations")
def upload_variations(request: Request, file: UploadFile = File(...)):
    email = require_uploader(request)
    raw = file.file.read()
    head = raw[:4096].decode("utf-8-sig", "replace")
    if "ASIN" not in head or "Variation ASINs" not in head:
        raise HTTPException(422, "Doesn't look like the variation export "
                                 "(needs ASIN + 'Variation ASINs' columns).")
    sha = commit_files({"data/master/keepa_variations.csv": raw},
                       "data(master): keepa variations upload via dashboard [skip ci]",
                       email)
    # Local mirror only after the commit succeeded.
    VARIATIONS_CSV.parent.mkdir(parents=True, exist_ok=True)
    VARIATIONS_CSV.write_bytes(raw)
    return {"ok": True, "commit": sha,
            "note": ("No change — identical to the current file." if sha is None
                     else "Committed. Variation Performance refreshes with the next "
                          "deploy (~5 minutes).")}
